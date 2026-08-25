from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from app.database import get_connection
from app.schemas import VehicleCreate, VehicleUpdate, VehicleClientTransfer, VehicleDeleteRequest
from app.security import get_current_user
from app.permissions import (
    can_create_request_for_client,
    can_open_client_details,
    has_any_permission,
)

import re
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/vehicles", tags=["Vehicles"])

VEHICLE_DELETE_ROLES = ["ADMIN"]

VEHICLE_TRASH_VIEW_ROLES = [
    "ADMIN",
    "ROP",
    "ACCOUNTANT",
    "MANAGER",
    "WAREHOUSE_MANAGER",
]

VEHICLE_RESTORE_ROLES = ["ADMIN"]

ALLOWED_VEHICLE_DELETE_REASON_TYPES = [
    "EQUIPMENT_REMOVED",
    "SERVICE_STOPPED_SIM_BLOCKED",
    "OTHER",
]

VEHICLE_IMPORT_REQUIRED_FIELDS = {
    "type": "Тип техники",
    "brand": "Марка",
    "model": "Модель",
    "vin": "VIN-код",
}

VEHICLE_IMPORT_HEADER_ALIASES = {
    "типтехники": "type",
    "тип": "type",

    "марка": "brand",
    "бренд": "brand",

    "модель": "model",

    "vin": "vin",
    "vinкод": "vin",
    "вин": "vin",
    "винкод": "vin",

    "госномер": "plate_number",
    "госномеравто": "plate_number",
    "госрегномер": "plate_number",
    "государственныйномер": "plate_number",
    "регномер": "plate_number",
    "номер": "plate_number",
    "номернойзнак": "plate_number",

    "годвыпуска": "year",
    "год": "year",
}


def clean_excel_cell(value) -> str:
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\xa0", " ")
    text = text.replace("\u200b", " ")

    return " ".join(text.strip().split())


def normalize_excel_header(value) -> str:
    text = clean_excel_cell(value).lower().replace("ё", "е")

    normalized = re.sub(r"[^a-zа-я0-9]+", "", text)

    # Чтобы заголовки из шаблона вроде:
    # "Гос. Номер (необязательно)"
    # "Год выпуска (необязательно)"
    # нормально превращались в "госномер" и "годвыпуска"
    optional_words = [
        "необязательно",
        "необязательное",
        "необязательный",
        "опционально",
    ]

    for word in optional_words:
        normalized = normalized.replace(word, "")

    return normalized


def normalize_vin(value) -> str:
    return re.sub(r"\s+", "", clean_excel_cell(value)).upper()


def normalize_plate_number(value) -> str | None:
    plate = re.sub(r"\s+", "", clean_excel_cell(value)).upper()

    return plate or None


def normalize_vehicle_text(value) -> str:
    return clean_excel_cell(value)


def parse_vehicle_year(value):
    text = clean_excel_cell(value)

    if not text:
        return None, None

    try:
        year = int(float(text))
    except ValueError:
        return None, f"Некорректный год выпуска: {text}"

    if year < 1900 or year > 2100:
        return None, f"Некорректный год выпуска: {text}"

    return year, None


def build_vehicle_import_header_map(header_row):
    header_map = {}

    for index, value in enumerate(header_row):
        normalized_header = normalize_excel_header(value)

        if not normalized_header:
            continue

        field_name = VEHICLE_IMPORT_HEADER_ALIASES.get(normalized_header)

        if not field_name:
            continue

        if field_name not in header_map:
            header_map[field_name] = index

    missing_fields = [
        label
        for field_name, label in VEHICLE_IMPORT_REQUIRED_FIELDS.items()
        if field_name not in header_map
    ]

    if missing_fields:
        raise HTTPException(
            status_code=400,
            detail="В Excel-файле не найдены обязательные колонки: "
            + ", ".join(missing_fields)
        )

    return header_map


def get_import_row_value(row, header_map, field_name):
    index = header_map.get(field_name)

    if index is None or index >= len(row):
        return None

    return row[index]

def get_client_display_name(client: dict) -> str:
    if not client:
        return "Клиент не найден"

    return (
        client.get("company_name")
        or client.get("name")
        or f"ID клиента {client.get('id')}"
    )


# ============================================================================
# Permission-aware access helpers
# ----------------------------------------------------------------------------
# Старые роли оставляем как fallback, чтобы текущие сотрудники не потеряли доступ
# после перехода на новую систему permissions.
# ============================================================================

VEHICLE_CREATE_PERMISSION_CODES = [
    "vehicles.create",
    "vehicles.manage",
]

VEHICLE_IMPORT_PERMISSION_CODES = [
    "vehicles.import",
    "vehicles.create",
    "vehicles.manage",
]

VEHICLE_VIEW_PERMISSION_CODES = [
    "vehicles.view",
    "vehicles.view_all",
    "vehicles.view_own",
    "vehicles.manage",
    "clients.view",
    "clients.view_all",
    "clients.view_own",
    "clients.manage",
]

VEHICLE_VIEW_ALL_PERMISSION_CODES = [
    "vehicles.view_all",
    "vehicles.manage",
    "clients.view_all",
    "clients.manage",
]

VEHICLE_EDIT_PERMISSION_CODES = [
    "vehicles.edit",
    "vehicles.edit_all",
    "vehicles.manage",
]

VEHICLE_EDIT_OWN_PERMISSION_CODES = [
    "vehicles.edit_own",
    "vehicles.manage_own",
]

VEHICLE_TRANSFER_PERMISSION_CODES = [
    "vehicles.transfer",
    "vehicles.transfer_client",
    "vehicles.manage",
]

VEHICLE_TRANSFER_HISTORY_PERMISSION_CODES = [
    "vehicles.transfer_history.view",
    "vehicles.transfer.view_history",
    "vehicles.transfer",
    "vehicles.transfer_client",
    "vehicles.manage",
]

VEHICLE_TRASH_VIEW_PERMISSION_CODES = [
    "vehicles.trash.view",
    "vehicles.deleted.view",
    "vehicles.restore",
    "vehicles.delete",
    "vehicles.manage",
    "trash.view",
    "trash.manage",
]

VEHICLE_DELETE_PERMISSION_CODES = [
    "vehicles.delete",
    "vehicles.manage",
]

VEHICLE_RESTORE_PERMISSION_CODES = [
    "vehicles.restore",
    "vehicles.manage",
]

VEHICLE_VIN_HISTORY_PERMISSION_CODES = [
    "vehicles.vin_history.view",
    "vehicles.history.view",
    "vehicles.view",
    "vehicles.view_all",
    "vehicles.manage",
]

VEHICLE_EQUIPMENT_MANAGE_PERMISSION_CODES = [
    "vehicles.equipment.manage",
    "warehouse.vehicle_equipment.manage",
    "warehouse.manage",
    "warehouse.items.manage",
]

VEHICLE_CREATE_LEGACY_ROLES = ["ADMIN", "ROP", "MANAGER", "TECH_SUPPORT"]
VEHICLE_IMPORT_LEGACY_ROLES = ["ADMIN", "ROP", "MANAGER", "TECH_SUPPORT"]
VEHICLE_VIEW_LEGACY_ROLES = [
    "ADMIN",
    "ROP",
    "MANAGER",
    "TECH_SUPPORT",
    "ACCOUNTANT",
    "WAREHOUSE_MANAGER",
    "SENIOR_TECHNICIAN",
    "TECHNICIAN",
]
VEHICLE_EDIT_LEGACY_ROLES = ["ADMIN", "ROP", "MANAGER"]
VEHICLE_TRANSFER_LEGACY_ROLES = ["ADMIN", "ROP", "MANAGER"]
VEHICLE_TRANSFER_HISTORY_LEGACY_ROLES = ["ADMIN", "ROP", "MANAGER"]


def permissions_are_loaded(current_user: dict | None) -> bool:
    return current_user is not None and isinstance(current_user.get("permissions"), list)


def has_legacy_role(current_user: dict | None, roles: list[str]) -> bool:
    if not current_user or permissions_are_loaded(current_user):
        return False

    return current_user.get("role") in roles


def can_create_vehicle(current_user: dict) -> bool:
    return has_any_permission(current_user, VEHICLE_CREATE_PERMISSION_CODES) or has_legacy_role(
        current_user,
        VEHICLE_CREATE_LEGACY_ROLES,
    )


def can_import_vehicles(current_user: dict) -> bool:
    return has_any_permission(current_user, VEHICLE_IMPORT_PERMISSION_CODES) or has_legacy_role(
        current_user,
        VEHICLE_IMPORT_LEGACY_ROLES,
    )


def can_search_vehicles(current_user: dict) -> bool:
    return has_any_permission(current_user, VEHICLE_VIEW_PERMISSION_CODES) or has_legacy_role(
        current_user,
        VEHICLE_VIEW_LEGACY_ROLES,
    )


def can_view_vehicle_trash(current_user: dict) -> bool:
    return has_any_permission(current_user, VEHICLE_TRASH_VIEW_PERMISSION_CODES) or has_legacy_role(
        current_user,
        VEHICLE_TRASH_VIEW_ROLES,
    )


def can_delete_vehicle(current_user: dict) -> bool:
    return has_any_permission(current_user, VEHICLE_DELETE_PERMISSION_CODES) or has_legacy_role(
        current_user,
        VEHICLE_DELETE_ROLES,
    )


def can_restore_vehicle(current_user: dict) -> bool:
    return has_any_permission(current_user, VEHICLE_RESTORE_PERMISSION_CODES) or has_legacy_role(
        current_user,
        VEHICLE_RESTORE_ROLES,
    )


def can_manage_direct_vehicle_equipment(current_user: dict) -> bool:
    return has_any_permission(current_user, VEHICLE_EQUIPMENT_MANAGE_PERMISSION_CODES) or has_legacy_role(
        current_user,
        ["ADMIN", "WAREHOUSE_MANAGER"],
    )


def build_client_from_vehicle_row(row: dict) -> dict:
    return {
        "id": row.get("client_id") or row.get("current_client_id"),
        "type": row.get("client_type") or row.get("current_client_type"),
        "name": row.get("client_name") or row.get("current_client_name"),
        "company_name": row.get("client_company_name") or row.get("current_client_company_name"),
        "status": row.get("client_status") or row.get("current_client_status"),
        "created_by": row.get("client_created_by") or row.get("created_by"),
        "responsible_manager_id": row.get("client_responsible_manager_id") or row.get("responsible_manager_id"),
        "is_deleted": row.get("client_is_deleted") or row.get("current_client_is_deleted"),
    }


def can_access_client_vehicles(client: dict, current_user: dict) -> bool:
    if has_any_permission(current_user, VEHICLE_VIEW_ALL_PERMISSION_CODES):
        return True

    if can_open_client_details(client, current_user):
        return True

    return False


def ensure_can_access_client_vehicles(client: dict, current_user: dict):
    if not can_access_client_vehicles(client, current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для просмотра машин этого клиента",
        )


def can_edit_vehicle_for_client(client: dict, current_user: dict) -> bool:
    if has_any_permission(current_user, VEHICLE_EDIT_PERMISSION_CODES):
        return True

    if has_legacy_role(current_user, ["ADMIN", "ROP"]):
        return True

    if has_any_permission(current_user, VEHICLE_EDIT_OWN_PERMISSION_CODES) or has_legacy_role(
        current_user,
        ["MANAGER"],
    ):
        return can_open_client_details(client, current_user)

    return False


def can_transfer_vehicle_for_client(client: dict, current_user: dict) -> bool:
    if has_any_permission(current_user, VEHICLE_TRANSFER_PERMISSION_CODES):
        return True

    if has_legacy_role(current_user, ["ADMIN", "ROP"]):
        return True

    if has_legacy_role(current_user, ["MANAGER"]):
        return can_create_request_for_client(client, current_user)

    return False


def can_view_vehicle_transfer_history_for_client(client: dict, current_user: dict) -> bool:
    if has_any_permission(current_user, VEHICLE_TRANSFER_HISTORY_PERMISSION_CODES):
        return True

    if has_legacy_role(current_user, VEHICLE_TRANSFER_HISTORY_LEGACY_ROLES):
        return can_open_client_details(client, current_user)

    return False


def can_view_vehicle_vin_history_for_client(client: dict, current_user: dict) -> bool:
    if has_any_permission(current_user, VEHICLE_VIN_HISTORY_PERMISSION_CODES):
        return True

    return can_open_client_details(client, current_user)

@router.post("")
def create_vehicle(data: VehicleCreate, current_user: dict = Depends(get_current_user)):
    if not can_create_vehicle(current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для создания машины"
        )

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            # 1. Проверяем клиента ДО проверки VIN и ДО создания машины.
            cursor.execute(
                """
                SELECT
                    id,
                    type,
                    name,
                    company_name,
                    status,
                    created_by,
                    responsible_manager_id,
                    is_deleted
                FROM clients
                WHERE id = %s
                LIMIT 1
                """,
                (data.client_id,)
            )

            client = cursor.fetchone()

            if not client or client["is_deleted"]:
                raise HTTPException(
                    status_code=404,
                    detail="Клиент не найден"
                )

            if not can_create_request_for_client(client, current_user):
                raise HTTPException(
                    status_code=403,
                    detail="Недостаточно прав для добавления машины этому клиенту"
                )

            if client.get("status") == "BLOCKED":
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя добавить машину заблокированному клиенту"
                )

            # 2. Только после проверки прав нормализуем и проверяем VIN.
            vin = normalize_vin(data.vin)

            if not vin:
                raise HTTPException(
                    status_code=400,
                    detail="VIN обязателен при создании машины"
                )

            cursor.execute(
                """
                SELECT
                    v.id,
                    v.client_id,
                    v.brand,
                    v.model,
                    v.plate_number,
                    c.name AS client_name,
                    c.company_name AS client_company_name
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                WHERE v.vin = %s
                  AND v.is_deleted = 0
                LIMIT 1
                """,
                (vin,)
            )

            existing_vehicle = cursor.fetchone()

            if existing_vehicle:
                client_name = (
                    existing_vehicle.get("client_company_name")
                    or existing_vehicle.get("client_name")
                    or f"ID клиента {existing_vehicle.get('client_id')}"
                )

                raise HTTPException(
                    status_code=400,
                    detail=f"Автомобиль с VIN {vin} уже существует у клиента: {client_name}"
                )

            cursor.execute(
                """
                SELECT
                    id,
                    client_id
                FROM vehicles
                WHERE vin = %s
                AND is_deleted = 1
                ORDER BY deleted_at DESC, id DESC
                LIMIT 1
                """,
                (vin,)
            )

            previous_deleted_vehicle_with_same_vin = cursor.fetchone()

            sql = """
            INSERT INTO vehicles (client_id, brand, model, plate_number, vin, year, type)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """

            cursor.execute(
                sql,
                (
                    data.client_id,
                    normalize_vehicle_text(data.brand),
                    normalize_vehicle_text(data.model),
                    normalize_plate_number(data.plate_number),
                    vin,
                    data.year,
                    normalize_vehicle_text(data.type),
                )
            )

            new_vehicle_id = cursor.lastrowid

            if previous_deleted_vehicle_with_same_vin:
                cursor.execute(
                    """
                    INSERT IGNORE INTO vehicle_vin_links (
                        vin,
                        old_vehicle_id,
                        new_vehicle_id,
                        old_client_id,
                        new_client_id,
                        created_by,
                        created_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, NOW())
                    """,
                    (
                        vin,
                        previous_deleted_vehicle_with_same_vin["id"],
                        new_vehicle_id,
                        previous_deleted_vehicle_with_same_vin.get("client_id"),
                        data.client_id,
                        current_user["id"],
                    )
                )

            connection.commit()

            return {
                "message": "created",
                "vehicle_id": new_vehicle_id,
                "linked_deleted_vehicles_count": 1 if previous_deleted_vehicle_with_same_vin else 0,
            }
        
    except HTTPException:
        connection.rollback()
        raise
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@router.get("")
def get_vehicles(
    client_id: int,
    limit: int | None = Query(default=None, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    current_user: dict = Depends(get_current_user),
):
    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    id,
                    type,
                    name,
                    company_name,
                    status,
                    created_by,
                    responsible_manager_id,
                    is_deleted
                FROM clients
                WHERE id = %s
                LIMIT 1
                """,
                (client_id,)
            )

            client = cursor.fetchone()

            if not client or client["is_deleted"]:
                raise HTTPException(status_code=404, detail="Клиент не найден")

            ensure_can_access_client_vehicles(client, current_user)

            cursor.execute(
                """
                SELECT COUNT(*) AS total
                FROM vehicles
                WHERE client_id = %s
                AND is_deleted = 0
                """,
                (client_id,)
            )
            total_row = cursor.fetchone()
            total = int(total_row["total"] or 0)

            if limit is None:
                cursor.execute(
                    """
                    SELECT *
                    FROM vehicles
                    WHERE client_id = %s
                    AND is_deleted = 0
                    ORDER BY id DESC
                    """,
                    (client_id,)
                )

                return cursor.fetchall()

            cursor.execute(
                """
                SELECT *
                FROM vehicles
                WHERE client_id = %s
                AND is_deleted = 0
                ORDER BY id DESC
                LIMIT %s OFFSET %s
                """,
                (client_id, limit, offset)
            )

            items = cursor.fetchall()

            return {
                "items": items,
                "total": total,
                "limit": limit,
                "offset": offset,
            }

    finally:
        connection.close()

@router.get("/search")
def search_vehicles(
    q: str = Query(..., min_length=2),
    limit: int = Query(default=10, ge=1, le=50),
    current_user: dict = Depends(get_current_user),
):
    """
    Быстрый глобальный поиск автомобилей без загрузки всех машин на фронт.
    Ищет по гос. номеру, VIN, марке, модели, клиенту, компании, телефону, БИН/ИИН.
    Для ролей с доступом к корзине также возвращает удалённые машины.
    """
    search = q.strip()

    if len(search) < 2:
        return []

    if not can_search_vehicles(current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для поиска машин"
        )

    like_value = f"%{search}%"
    vin_search = normalize_vin(search)

    conditions = [
        "c.is_deleted = 0",
        """
        (
            v.plate_number LIKE %s OR
            v.vin LIKE %s OR
            v.brand LIKE %s OR
            v.model LIKE %s OR
            c.name LIKE %s OR
            c.company_name LIKE %s OR
            c.phone LIKE %s OR
            c.bin_iin LIKE %s
        )
        """,
    ]

    values = [
        like_value,
        like_value,
        like_value,
        like_value,
        like_value,
        like_value,
        like_value,
        like_value,
    ]

    if not can_view_vehicle_trash(current_user):
        conditions.append("v.is_deleted = 0")

    if current_user.get("client_access_scope") == "RESPONSIBLE_ONLY":
        conditions.append("c.responsible_manager_id = %s")
        values.append(current_user["id"])

    where_clause = " AND ".join(conditions)

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            sql = f"""
            SELECT
                v.id,
                v.client_id,
                v.brand,
                v.model,
                v.plate_number,
                v.vin,
                v.year,
                v.type,
                v.is_deleted,
                v.deleted_at,
                v.deleted_by,
                v.delete_reason_type,
                v.delete_reason,

                c.name AS client_name,
                c.company_name,
                c.phone AS client_phone,
                c.bin_iin AS client_bin_iin,
                c.type AS client_type,
                c.status AS client_status,
                c.responsible_manager_id,

                responsible.name AS responsible_manager_name,

                deleted_by_user.name AS deleted_by_name
            FROM vehicles v
            LEFT JOIN clients c ON v.client_id = c.id
            LEFT JOIN users responsible ON c.responsible_manager_id = responsible.id
            LEFT JOIN users deleted_by_user ON v.deleted_by = deleted_by_user.id
            WHERE {where_clause}
            ORDER BY
                CASE
                    WHEN v.is_deleted = 0 THEN 0
                    ELSE 1
                END,
                CASE
                    WHEN v.vin = %s THEN 1
                    WHEN v.plate_number = %s THEN 2
                    WHEN v.vin LIKE %s THEN 3
                    WHEN v.plate_number LIKE %s THEN 4
                    WHEN c.bin_iin = %s THEN 5
                    WHEN c.phone = %s THEN 6
                    WHEN c.company_name LIKE %s THEN 7
                    WHEN c.name LIKE %s THEN 8
                    ELSE 9
                END,
                v.is_deleted ASC,
                v.deleted_at DESC,
                v.id DESC
            LIMIT %s
            """

            values.extend([
                vin_search,
                search,
                like_value,
                like_value,
                search,
                search,
                like_value,
                like_value,
                limit,
            ])

            cursor.execute(sql, tuple(values))
            return cursor.fetchall()

    finally:
        connection.close()

@router.get("/check-vin")
def check_vehicle_vin(vin: str, current_user: dict = Depends(get_current_user)):
    """
    Проверка VIN перед созданием автомобиля.
    Нужна, чтобы фронт мог проверить VIN до создания нового клиента.
    """
    if not can_create_vehicle(current_user) and not can_import_vehicles(current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для проверки VIN"
        )

    normalized_vin = normalize_vin(vin)

    if not normalized_vin:
        return {
            "exists": False,
            "vin": None,
            "vehicle": None
        }

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    v.id,
                    v.client_id,
                    v.brand,
                    v.model,
                    v.plate_number,
                    v.vin,
                    v.is_deleted AS vehicle_is_deleted,

                    c.name AS client_name,
                    c.company_name,
                    c.type AS client_type,
                    c.is_deleted AS client_is_deleted
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                WHERE v.vin = %s
                AND v.is_deleted = 0
                LIMIT 1
                """,
                (normalized_vin,)
            )

            vehicle = cursor.fetchone()

            return {
                "exists": vehicle is not None,
                "vin": normalized_vin,
                "vehicle": vehicle
            }

    finally:
        connection.close()

@router.get("/{vehicle_id}/vin-history")
def get_vehicle_vin_history(
    vehicle_id: int,
    current_user: dict = Depends(get_current_user),
):
    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    v.id,
                    v.client_id,
                    v.brand,
                    v.model,
                    v.plate_number,
                    v.vin,
                    v.year,
                    v.type,
                    v.is_deleted,
                    v.deleted_at,
                    v.deleted_by,
                    v.delete_reason_type,
                    v.delete_reason,

                    c.name AS client_name,
                    c.company_name AS client_company_name,
                    c.type AS client_type,
                    c.status AS client_status,
                    c.created_by AS client_created_by,
                    c.responsible_manager_id AS client_responsible_manager_id,
                    c.is_deleted AS client_is_deleted,

                    deleted_by_user.name AS deleted_by_name
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                LEFT JOIN users deleted_by_user ON v.deleted_by = deleted_by_user.id
                WHERE v.id = %s
                LIMIT 1
                """,
                (vehicle_id,)
            )

            current_vehicle = cursor.fetchone()

            if not current_vehicle:
                raise HTTPException(
                    status_code=404,
                    detail="Машина не найдена"
                )

            current_client = build_client_from_vehicle_row(current_vehicle)

            if not can_view_vehicle_vin_history_for_client(current_client, current_user):
                raise HTTPException(
                    status_code=403,
                    detail="Недостаточно прав для просмотра истории VIN этой машины"
                )

            vin = normalize_vin(current_vehicle.get("vin"))

            if not vin:
                return {
                    "vehicle": current_vehicle,
                    "vin": None,
                    "related_vehicles": [],
                    "links": [],
                    "links_count": 0,
                }

            cursor.execute(
                """
                SELECT
                    v.id,
                    v.client_id,
                    v.brand,
                    v.model,
                    v.plate_number,
                    v.vin,
                    v.year,
                    v.type,
                    v.is_deleted,
                    v.deleted_at,
                    v.deleted_by,
                    v.delete_reason_type,
                    v.delete_reason,

                    c.name AS client_name,
                    c.company_name AS client_company_name,

                    deleted_by_user.name AS deleted_by_name
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                LEFT JOIN users deleted_by_user ON v.deleted_by = deleted_by_user.id
                WHERE v.vin = %s
                ORDER BY
                    v.is_deleted ASC,
                    v.deleted_at DESC,
                    v.id DESC
                """,
                (vin,)
            )

            related_vehicles = cursor.fetchall()

            cursor.execute(
                """
                SELECT
                    l.id,
                    l.vin,
                    l.old_vehicle_id,
                    l.new_vehicle_id,
                    l.old_client_id,
                    l.new_client_id,
                    l.created_by,
                    l.created_at,

                    creator.name AS created_by_name,

                    old_v.brand AS old_brand,
                    old_v.model AS old_model,
                    old_v.plate_number AS old_plate_number,
                    old_v.vin AS old_vin,
                    old_v.year AS old_year,
                    old_v.type AS old_type,
                    old_v.is_deleted AS old_is_deleted,
                    old_v.deleted_at AS old_deleted_at,
                    old_v.delete_reason_type AS old_delete_reason_type,
                    old_v.delete_reason AS old_delete_reason,

                    old_client.name AS old_client_name,
                    old_client.company_name AS old_client_company_name,

                    old_deleted_by.name AS old_deleted_by_name,

                    new_v.brand AS new_brand,
                    new_v.model AS new_model,
                    new_v.plate_number AS new_plate_number,
                    new_v.vin AS new_vin,
                    new_v.year AS new_year,
                    new_v.type AS new_type,
                    new_v.is_deleted AS new_is_deleted,
                    new_v.deleted_at AS new_deleted_at,
                    new_v.delete_reason_type AS new_delete_reason_type,
                    new_v.delete_reason AS new_delete_reason,

                    new_client.name AS new_client_name,
                    new_client.company_name AS new_client_company_name,

                    new_deleted_by.name AS new_deleted_by_name
                FROM vehicle_vin_links l
                LEFT JOIN users creator ON l.created_by = creator.id

                LEFT JOIN vehicles old_v ON l.old_vehicle_id = old_v.id
                LEFT JOIN clients old_client ON l.old_client_id = old_client.id
                LEFT JOIN users old_deleted_by ON old_v.deleted_by = old_deleted_by.id

                LEFT JOIN vehicles new_v ON l.new_vehicle_id = new_v.id
                LEFT JOIN clients new_client ON l.new_client_id = new_client.id
                LEFT JOIN users new_deleted_by ON new_v.deleted_by = new_deleted_by.id

                WHERE l.vin = %s
                   OR l.old_vehicle_id = %s
                   OR l.new_vehicle_id = %s
                ORDER BY l.created_at DESC, l.id DESC
                """,
                (
                    vin,
                    vehicle_id,
                    vehicle_id,
                )
            )

            link_rows = cursor.fetchall()

            links = []

            for row in link_rows:
                links.append({
                    "id": row["id"],
                    "vin": row["vin"],
                    "created_at": row["created_at"],
                    "created_by": row["created_by"],
                    "created_by_name": row["created_by_name"],

                    "old_vehicle": {
                        "id": row["old_vehicle_id"],
                        "client_id": row["old_client_id"],
                        "brand": row["old_brand"],
                        "model": row["old_model"],
                        "plate_number": row["old_plate_number"],
                        "vin": row["old_vin"],
                        "year": row["old_year"],
                        "type": row["old_type"],
                        "is_deleted": row["old_is_deleted"],
                        "deleted_at": row["old_deleted_at"],
                        "delete_reason_type": row["old_delete_reason_type"],
                        "delete_reason": row["old_delete_reason"],
                        "deleted_by_name": row["old_deleted_by_name"],
                        "client_name": row["old_client_name"],
                        "client_company_name": row["old_client_company_name"],
                    },

                    "new_vehicle": {
                        "id": row["new_vehicle_id"],
                        "client_id": row["new_client_id"],
                        "brand": row["new_brand"],
                        "model": row["new_model"],
                        "plate_number": row["new_plate_number"],
                        "vin": row["new_vin"],
                        "year": row["new_year"],
                        "type": row["new_type"],
                        "is_deleted": row["new_is_deleted"],
                        "deleted_at": row["new_deleted_at"],
                        "delete_reason_type": row["new_delete_reason_type"],
                        "delete_reason": row["new_delete_reason"],
                        "deleted_by_name": row["new_deleted_by_name"],
                        "client_name": row["new_client_name"],
                        "client_company_name": row["new_client_company_name"],
                    },
                })

            return {
                "vehicle": current_vehicle,
                "vin": vin,
                "related_vehicles": related_vehicles,
                "links": links,
                "links_count": len(links),
            }

    finally:
        connection.close()

@router.get("/import-template")
def download_vehicle_import_template(
    current_user: dict = Depends(get_current_user),
):
    if not can_import_vehicles(current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для скачивания шаблона импорта машин"
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Автомобили"

    headers = [
        "Тип техники",
        "Марка",
        "Модель",
        "VIN-код",
        "Гос. Номер (необязательно)",
        "Год выпуска (необязательно)",
    ]

    example_rows = [
        ["Легковая", "Toyota", "Camry", "VINTOYOTACAMRY123", "", ""],
        ["Электромобиль", "BYD", "Han", "VINBYDHAN123123123", "", ""],
        ["Спецтехника", "KAMAZ", "4310", "VINKAMAZ4310123123", "", ""],
    ]

    sheet.append(headers)

    for row in example_rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="5E9424")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 30
    sheet.column_dimensions["F"].width = 30

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="vehicle_import_template.xlsx"'
        },
    )

@router.post("/import-preview")
def import_vehicles_preview(
    client_id: int | None = Form(None),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Preview импорта авто из .xlsx для создания заявки.

    Ничего не создаёт в базе.
    Возвращает:
    - existing: VIN найден у выбранного клиента
    - new: VIN не найден в CRM
    - warnings: строка пропущена
    """
    if not can_import_vehicles(current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для импорта машин"
        )

    filename = file.filename or ""

    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Можно импортировать только Excel-файл формата .xlsx"
        )

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            selected_client_id = int(client_id) if client_id else None
            client = None

            if selected_client_id:
                cursor.execute(
                    """
                    SELECT
                        id,
                        type,
                        name,
                        company_name,
                        status,
                        created_by,
                        responsible_manager_id,
                        is_deleted
                    FROM clients
                    WHERE id = %s
                    LIMIT 1
                    """,
                    (selected_client_id,)
                )

                client = cursor.fetchone()

                if not client or client["is_deleted"]:
                    raise HTTPException(
                        status_code=404,
                        detail="Клиент не найден"
                    )

                if client.get("status") == "BLOCKED":
                    raise HTTPException(
                        status_code=400,
                        detail="Нельзя импортировать машины для заблокированного клиента"
                    )

                if not can_create_request_for_client(client, current_user):
                    raise HTTPException(
                        status_code=403,
                        detail="Недостаточно прав для импорта машин этому клиенту"
                    )

            content = file.file.read()

            if not content:
                raise HTTPException(
                    status_code=400,
                    detail="Excel-файл пустой"
                )

            if len(content) > 5 * 1024 * 1024:
                raise HTTPException(
                    status_code=400,
                    detail="Excel-файл слишком большой. Максимум 5 МБ"
                )

            try:
                workbook = load_workbook(
                    filename=BytesIO(content),
                    read_only=True,
                    data_only=True,
                )
            except Exception:
                raise HTTPException(
                    status_code=400,
                    detail="Не удалось прочитать Excel-файл. Проверьте, что это настоящий .xlsx файл"
                )

            sheet = workbook.active

            rows_iterator = sheet.iter_rows(values_only=True)

            try:
                header_row = next(rows_iterator)
            except StopIteration:
                raise HTTPException(
                    status_code=400,
                    detail="Excel-файл пустой"
                )

            header_map = build_vehicle_import_header_map(header_row)

            parsed_rows = []
            warnings = []
            seen_vins = set()
            total_rows = 0

            for row_number, row in enumerate(rows_iterator, start=2):
                if row_number > 1002:
                    warnings.append({
                        "row": row_number,
                        "vin": None,
                        "message": "Строка пропущена: за один импорт можно загрузить максимум 1000 машин"
                    })
                    continue

                row_has_value = any(clean_excel_cell(value) for value in row)

                if not row_has_value:
                    continue

                total_rows += 1

                vehicle_type = normalize_vehicle_text(
                    get_import_row_value(row, header_map, "type")
                )
                brand = normalize_vehicle_text(
                    get_import_row_value(row, header_map, "brand")
                )
                model = normalize_vehicle_text(
                    get_import_row_value(row, header_map, "model")
                )
                vin = normalize_vin(
                    get_import_row_value(row, header_map, "vin")
                )
                plate_number = normalize_plate_number(
                    get_import_row_value(row, header_map, "plate_number")
                )
                year, year_warning = parse_vehicle_year(
                    get_import_row_value(row, header_map, "year")
                )

                missing = []

                if not vehicle_type:
                    missing.append("Тип техники")

                if not brand:
                    missing.append("Марка")

                if not model:
                    missing.append("Модель")

                if not vin:
                    missing.append("VIN-код")

                if missing:
                    warnings.append({
                        "row": row_number,
                        "vin": vin or None,
                        "message": "Строка пропущена: не заполнены обязательные поля: "
                        + ", ".join(missing)
                    })
                    continue

                if year_warning:
                    warnings.append({
                        "row": row_number,
                        "vin": vin,
                        "message": year_warning + ". Год будет оставлен пустым"
                    })

                if vin in seen_vins:
                    warnings.append({
                        "row": row_number,
                        "vin": vin,
                        "message": f"Строка пропущена: VIN {vin} повторяется внутри Excel-файла"
                    })
                    continue

                seen_vins.add(vin)

                parsed_rows.append({
                    "row": row_number,
                    "type": vehicle_type,
                    "brand": brand,
                    "model": model,
                    "vin": vin,
                    "plate_number": plate_number,
                    "year": year,
                })

            if not parsed_rows:
                return {
                    "items": [],
                    "warnings": warnings,
                    "summary": {
                        "total_rows": total_rows,
                        "imported_count": 0,
                        "existing_count": 0,
                        "new_count": 0,
                        "skipped_count": len(warnings),
                    }
                }

            vins = [row["vin"] for row in parsed_rows]
            placeholders = ", ".join(["%s"] * len(vins))

            cursor.execute(
                f"""
                SELECT
                    v.id,
                    v.client_id,
                    v.brand,
                    v.model,
                    v.plate_number,
                    v.vin,
                    v.year,
                    v.type,

                    c.name AS client_name,
                    c.company_name AS client_company_name
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                WHERE v.vin IN ({placeholders})
                  AND v.is_deleted = 0
                """,
                tuple(vins)
            )

            existing_vehicles = cursor.fetchall()
            existing_by_vin = {}

            for vehicle in existing_vehicles:
                vehicle_vin = normalize_vin(vehicle.get("vin"))

                if vehicle_vin and vehicle_vin not in existing_by_vin:
                    existing_by_vin[vehicle_vin] = vehicle

            items = []
            existing_count = 0
            new_count = 0

            for row in parsed_rows:
                existing_vehicle = existing_by_vin.get(row["vin"])

                if not existing_vehicle:
                    new_count += 1

                    items.append({
                        "row": row["row"],
                        "mode": "new",
                        "vehicle_id": None,
                        "client_id": selected_client_id,
                        "type": row["type"],
                        "brand": row["brand"],
                        "model": row["model"],
                        "vin": row["vin"],
                        "plate_number": row["plate_number"],
                        "year": row["year"],
                    })
                    continue

                if (
                    selected_client_id
                    and int(existing_vehicle["client_id"]) == int(selected_client_id)
                ):
                    existing_count += 1

                    excel_plate_number = row.get("plate_number")
                    excel_year = row.get("year")

                    items.append({
                        "row": row["row"],
                        "mode": "existing",
                        "vehicle_id": existing_vehicle["id"],
                        "client_id": existing_vehicle["client_id"],
                        "type": existing_vehicle["type"],
                        "brand": existing_vehicle["brand"],
                        "model": existing_vehicle["model"],
                        "vin": existing_vehicle["vin"],

                        # В preview показываем значение из Excel, если оно заполнено.
                        # Если в Excel пусто — оставляем значение из базы.
                        "plate_number": excel_plate_number or existing_vehicle["plate_number"],
                        "year": excel_year if excel_year is not None else existing_vehicle["year"],

                        # Технические поля, чтобы frontend понимал:
                        # Excel предлагает обновить данные существующей машины.
                        "excel_plate_number": excel_plate_number,
                        "db_plate_number": existing_vehicle["plate_number"],
                        "excel_year": excel_year,
                        "db_year": existing_vehicle["year"],
                        "needs_vehicle_update": bool(
                            excel_plate_number
                            and excel_plate_number != existing_vehicle["plate_number"]
                        ) or (
                            excel_year is not None
                            and excel_year != existing_vehicle["year"]
                        ),
                    })
                    continue

                other_client_name = (
                    existing_vehicle.get("client_company_name")
                    or existing_vehicle.get("client_name")
                    or f"ID клиента {existing_vehicle.get('client_id')}"
                )

                warnings.append({
                    "row": row["row"],
                    "vin": row["vin"],
                    "message": f"VIN {row['vin']} уже привязан к другому клиенту: {other_client_name}. Машина пропущена."
                })

            return {
                "items": items,
                "warnings": warnings,
                "summary": {
                    "total_rows": total_rows,
                    "imported_count": len(items),
                    "existing_count": existing_count,
                    "new_count": new_count,
                    "skipped_count": len(warnings),
                }
            }

    finally:
        connection.close()

@router.get("/deleted")
def get_deleted_vehicles(
    client_id: int | None = Query(None),
    current_user: dict = Depends(get_current_user),
):
    """Список удалённых машин."""
    if not can_view_vehicle_trash(current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для просмотра корзины машин"
        )

    conditions = ["v.is_deleted = 1"]
    values = []

    if client_id:
        conditions.append("v.client_id = %s")
        values.append(client_id)

    if current_user.get("client_access_scope") == "RESPONSIBLE_ONLY":
        conditions.append("c.responsible_manager_id = %s")
        values.append(current_user["id"])

    where_clause = " AND ".join(conditions)

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            sql = f"""
            SELECT 
                v.id,
                v.client_id,
                v.brand,
                v.model,
                v.plate_number,
                v.vin,
                v.year,
                v.type,
                v.deleted_at,
                v.deleted_by,
                v.delete_reason_type,
                v.delete_reason,

                c.name AS client_name,
                c.company_name,
                c.type AS client_type,

                u.name AS deleted_by_name,

                COUNT(DISTINCT r.id) AS request_count
            FROM vehicles v
            LEFT JOIN clients c ON v.client_id = c.id
            LEFT JOIN users u ON v.deleted_by = u.id
            LEFT JOIN request_vehicles rv ON v.id = rv.vehicle_id
            LEFT JOIN requests r 
                ON rv.request_id = r.id
                AND r.is_deleted = 0
            WHERE {where_clause}
            GROUP BY 
                v.id,
                v.client_id,
                v.brand,
                v.model,
                v.plate_number,
                v.vin,
                v.year,
                v.type,
                v.deleted_at,
                v.deleted_by,
                v.delete_reason_type,
                v.delete_reason,
                c.name,
                c.company_name,
                c.type,
                u.name
            ORDER BY v.deleted_at DESC
            """

            cursor.execute(sql, tuple(values))
            return cursor.fetchall()

    finally:
        connection.close()

@router.get("/{vehicle_id}/page")
def get_vehicle_page(
    vehicle_id: int,
    limit: int = Query(default=20, ge=1, le=200),
    current_user: dict = Depends(get_current_user),
):
    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    v.id,
                    v.client_id,
                    v.brand,
                    v.model,
                    v.plate_number,
                    v.vin,
                    c.type AS client_type,
                    c.name AS client_name,
                    c.company_name AS client_company_name,
                    c.status AS client_status,
                    c.created_by AS client_created_by,
                    c.responsible_manager_id AS client_responsible_manager_id,
                    c.is_deleted AS client_is_deleted
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                WHERE v.id = %s
                  AND v.is_deleted = 0
                LIMIT 1
                """,
                (vehicle_id,)
            )

            vehicle = cursor.fetchone()

            if not vehicle:
                raise HTTPException(
                    status_code=404,
                    detail="Машина не найдена"
                )

            client = build_client_from_vehicle_row(vehicle)
            ensure_can_access_client_vehicles(client, current_user)

            client_id = vehicle["client_id"]

            cursor.execute(
                """
                SELECT COUNT(*) AS total
                FROM vehicles
                WHERE client_id = %s
                  AND is_deleted = 0
                """,
                (client_id,)
            )
            total_row = cursor.fetchone()
            total = int(total_row["total"] or 0)

            cursor.execute(
                """
                SELECT COUNT(*) AS before_count
                FROM vehicles
                WHERE client_id = %s
                  AND is_deleted = 0
                  AND id > %s
                """,
                (client_id, vehicle_id)
            )
            before_row = cursor.fetchone()
            before_count = int(before_row["before_count"] or 0)

            position = before_count + 1
            page = ((position - 1) // limit) + 1
            offset = (page - 1) * limit

            return {
                "vehicle": vehicle,
                "client_id": client_id,
                "vehicle_id": vehicle_id,
                "total": total,
                "position": position,
                "page": page,
                "limit": limit,
                "offset": offset,
            }

    finally:
        connection.close()

@router.patch("/{vehicle_id}")
def update_vehicle(
    vehicle_id: int,
    data: VehicleUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Редактирование машины."""
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT 
                    v.id,
                    v.client_id,
                    v.is_deleted,
                    c.type AS client_type,
                    c.name AS client_name,
                    c.company_name AS client_company_name,
                    c.status AS client_status,
                    c.created_by AS client_created_by,
                    c.responsible_manager_id AS client_responsible_manager_id,
                    c.is_deleted AS client_is_deleted
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                WHERE v.id = %s
                """,
                (vehicle_id,)
            )
            vehicle = cursor.fetchone()

            if not vehicle:
                raise HTTPException(status_code=404, detail="Машина не найдена")

            if vehicle["is_deleted"]:
                raise HTTPException(status_code=400, detail="Нельзя редактировать машину из корзины")

            if vehicle["client_id"] is None or vehicle["client_is_deleted"]:
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя редактировать машину: клиент не найден или находится в корзине"
                )

            client = build_client_from_vehicle_row(vehicle)

            if not can_edit_vehicle_for_client(client, current_user):
                raise HTTPException(
                    status_code=403,
                    detail="Недостаточно прав для редактирования этой машины"
                )

            update_data = data.dict(exclude_unset=True)

            if not update_data:
                return {"message": "Нет данных для обновления"}
            
            if "vin" in update_data:
                new_vin = normalize_vin(update_data["vin"])

                if not new_vin:
                    raise HTTPException(
                        status_code=400,
                        detail="VIN обязателен. Нельзя сохранить машину без VIN"
                    )

                cursor.execute(
                    """
                    SELECT id
                    FROM vehicles
                    WHERE vin = %s
                    AND id != %s
                    AND is_deleted = 0
                    """,
                    (new_vin, vehicle_id)
                )
                existing_vehicle = cursor.fetchone()

                if existing_vehicle:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Автомобиль с VIN {new_vin} уже существует"
                    )

                update_data["vin"] = new_vin

            allowed_fields = ["brand", "model", "plate_number", "vin", "year", "type"]

            updates = []
            values = []

            for text_field in ["brand", "model", "type"]:
                if text_field in update_data:
                    update_data[text_field] = normalize_vehicle_text(update_data[text_field])

            if "plate_number" in update_data:
                update_data["plate_number"] = normalize_plate_number(update_data["plate_number"])

            for field in allowed_fields:
                if field in update_data:
                    updates.append(f"{field} = %s")
                    values.append(update_data[field])

            if not updates:
                return {"message": "Нет допустимых полей для обновления"}

            values.append(vehicle_id)

            sql = f"""
            UPDATE vehicles
            SET {', '.join(updates)}
            WHERE id = %s
            """

            cursor.execute(sql, tuple(values))
            connection.commit()

            return {
                "message": "Машина обновлена",
                "vehicle_id": vehicle_id
            }

    except HTTPException:
        raise
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@router.post("/{vehicle_id}/transfer-client")
def transfer_vehicle_to_client(
    vehicle_id: int,
    data: VehicleClientTransfer,
    current_user: dict = Depends(get_current_user)
):
    """
    Перенос машины от одного клиента к другому.

    Меняем только vehicles.client_id.
    Старые заявки, request_vehicles и request_equipment не переносим,
    потому что они являются историей работ старого клиента.
    """
    if not has_any_permission(current_user, VEHICLE_TRANSFER_PERMISSION_CODES) and not has_legacy_role(
        current_user,
        VEHICLE_TRANSFER_LEGACY_ROLES,
    ):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для переноса машины между клиентами"
        )

    reason = (data.reason or "").strip()

    if not reason:
        raise HTTPException(
            status_code=400,
            detail="Необходимо указать причину переноса"
        )

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    v.id,
                    v.client_id,
                    v.brand,
                    v.model,
                    v.plate_number,
                    v.vin,
                    v.is_deleted,

                    old_client.id AS old_client_id,
                    old_client.type AS old_client_type,
                    old_client.name AS old_client_name,
                    old_client.company_name AS old_client_company_name,
                    old_client.status AS old_client_status,
                    old_client.created_by AS old_client_created_by,
                    old_client.responsible_manager_id AS old_client_responsible_manager_id,
                    old_client.is_deleted AS old_client_is_deleted
                FROM vehicles v
                LEFT JOIN clients old_client ON v.client_id = old_client.id
                WHERE v.id = %s
                LIMIT 1
                """,
                (vehicle_id,)
            )

            vehicle = cursor.fetchone()

            if not vehicle:
                raise HTTPException(
                    status_code=404,
                    detail="Машина не найдена"
                )

            if vehicle["is_deleted"]:
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя переносить машину из корзины"
                )

            if not vehicle["old_client_id"] or vehicle["old_client_is_deleted"]:
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя перенести машину: текущий клиент не найден или находится в корзине"
                )

            old_client = {
                "id": vehicle["old_client_id"],
                "type": vehicle["old_client_type"],
                "name": vehicle["old_client_name"],
                "company_name": vehicle["old_client_company_name"],
                "status": vehicle["old_client_status"],
                "created_by": vehicle["old_client_created_by"],
                "responsible_manager_id": vehicle["old_client_responsible_manager_id"],
                "is_deleted": vehicle["old_client_is_deleted"],
            }

            if not can_transfer_vehicle_for_client(old_client, current_user):
                raise HTTPException(
                    status_code=403,
                    detail="Недостаточно прав для переноса машины от текущего клиента"
                )

            if int(vehicle["client_id"]) == int(data.new_client_id):
                raise HTTPException(
                    status_code=400,
                    detail="Машина уже находится у выбранного клиента"
                )

            cursor.execute(
                """
                SELECT
                    id,
                    type,
                    name,
                    company_name,
                    status,
                    created_by,
                    responsible_manager_id,
                    is_deleted
                FROM clients
                WHERE id = %s
                LIMIT 1
                """,
                (data.new_client_id,)
            )

            new_client = cursor.fetchone()

            if not new_client or new_client["is_deleted"]:
                raise HTTPException(
                    status_code=404,
                    detail="Новый клиент не найден"
                )

            if not can_transfer_vehicle_for_client(new_client, current_user):
                raise HTTPException(
                    status_code=403,
                    detail="Недостаточно прав для переноса машины к выбранному клиенту"
                )

            if new_client.get("status") == "BLOCKED":
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя перенести машину к заблокированному клиенту"
                )

            cursor.execute(
                """
                SELECT COUNT(DISTINCT rv.request_id) AS request_count
                FROM request_vehicles rv
                INNER JOIN requests r ON rv.request_id = r.id
                WHERE rv.vehicle_id = %s
                  AND r.is_deleted = 0
                """,
                (vehicle_id,)
            )
            request_count_row = cursor.fetchone()
            request_count = int(request_count_row["request_count"] or 0)

            cursor.execute(
                """
                SELECT COUNT(re.id) AS equipment_count
                FROM request_equipment re
                INNER JOIN request_vehicles rv ON re.request_vehicle_id = rv.id
                INNER JOIN requests r ON rv.request_id = r.id
                WHERE rv.vehicle_id = %s
                  AND r.is_deleted = 0
                """,
                (vehicle_id,)
            )
            equipment_count_row = cursor.fetchone()
            equipment_count = int(equipment_count_row["equipment_count"] or 0)

            old_client_id = int(vehicle["client_id"])
            new_client_id = int(data.new_client_id)

            cursor.execute(
                """
                UPDATE vehicles
                SET client_id = %s
                WHERE id = %s
                """,
                (new_client_id, vehicle_id)
            )

            cursor.execute(
                """
                INSERT INTO vehicle_transfer_history (
                    vehicle_id,
                    old_client_id,
                    new_client_id,
                    reason,
                    created_by
                )
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    vehicle_id,
                    old_client_id,
                    new_client_id,
                    reason,
                    current_user["id"],
                )
            )

            history_id = cursor.lastrowid

            connection.commit()

            return {
                "message": "Машина перенесена к другому клиенту",
                "vehicle_id": vehicle_id,
                "old_client_id": old_client_id,
                "old_client_name": get_client_display_name({
                    "id": old_client_id,
                    "name": vehicle["old_client_name"],
                    "company_name": vehicle["old_client_company_name"],
                }),
                "new_client_id": new_client_id,
                "new_client_name": get_client_display_name(new_client),
                "reason": reason,
                "history_id": history_id,
                "historical_requests_left_unchanged": request_count,
                "equipment_links_left_unchanged": equipment_count,
            }

    except HTTPException:
        connection.rollback()
        raise
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@router.get("/{vehicle_id}/transfer-history")
def get_vehicle_transfer_history(
    vehicle_id: int,
    current_user: dict = Depends(get_current_user)
):
    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    v.id,
                    v.client_id,
                    v.is_deleted,

                    c.id AS current_client_id,
                    c.type AS current_client_type,
                    c.name AS current_client_name,
                    c.company_name AS current_client_company_name,
                    c.status AS current_client_status,
                    c.created_by,
                    c.responsible_manager_id,
                    c.is_deleted AS current_client_is_deleted
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                WHERE v.id = %s
                LIMIT 1
                """,
                (vehicle_id,)
            )

            vehicle = cursor.fetchone()

            if not vehicle:
                raise HTTPException(
                    status_code=404,
                    detail="Машина не найдена"
                )

            if vehicle["is_deleted"]:
                raise HTTPException(
                    status_code=400,
                    detail="Машина находится в корзине"
                )

            current_client = {
                "id": vehicle["current_client_id"],
                "type": vehicle["current_client_type"],
                "name": vehicle["current_client_name"],
                "company_name": vehicle["current_client_company_name"],
                "status": vehicle.get("current_client_status"),
                "created_by": vehicle["created_by"],
                "responsible_manager_id": vehicle["responsible_manager_id"],
                "is_deleted": vehicle["current_client_is_deleted"],
            }

            if not can_view_vehicle_transfer_history_for_client(current_client, current_user):
                raise HTTPException(
                    status_code=403,
                    detail="Недостаточно прав для просмотра истории переноса этой машины"
                )

            cursor.execute(
                """
                SELECT
                    h.id,
                    h.vehicle_id,
                    h.old_client_id,
                    old_client.type AS old_client_type,
                    old_client.name AS old_client_name,
                    old_client.company_name AS old_client_company_name,

                    h.new_client_id,
                    new_client.type AS new_client_type,
                    new_client.name AS new_client_name,
                    new_client.company_name AS new_client_company_name,

                    h.reason,
                    h.created_by,
                    creator.name AS created_by_name,
                    h.created_at
                FROM vehicle_transfer_history h
                LEFT JOIN clients old_client ON h.old_client_id = old_client.id
                LEFT JOIN clients new_client ON h.new_client_id = new_client.id
                LEFT JOIN users creator ON h.created_by = creator.id
                WHERE h.vehicle_id = %s
                ORDER BY h.created_at DESC, h.id DESC
                """,
                (vehicle_id,)
            )

            rows = cursor.fetchall()

            for row in rows:
                row["old_client_display_name"] = get_client_display_name({
                    "id": row["old_client_id"],
                    "name": row["old_client_name"],
                    "company_name": row["old_client_company_name"],
                })

                row["new_client_display_name"] = get_client_display_name({
                    "id": row["new_client_id"],
                    "name": row["new_client_name"],
                    "company_name": row["new_client_company_name"],
                })

            return rows

    finally:
        connection.close()

@router.delete("/{vehicle_id}")
def delete_vehicle(
    vehicle_id: int,
    data: VehicleDeleteRequest,
    current_user: dict = Depends(get_current_user)
):
    """Soft delete машины."""
    if not can_delete_vehicle(current_user):
        raise HTTPException(
            status_code=403,
            detail="Недостаточно прав для удаления машины"
        )

    delete_reason_type = (data.delete_reason_type or "").strip()
    delete_reason = (data.delete_reason or "").strip()

    if delete_reason_type not in ALLOWED_VEHICLE_DELETE_REASON_TYPES:
        raise HTTPException(
            status_code=400,
            detail="Некорректный тип причины удаления машины"
        )

    if not delete_reason:
        raise HTTPException(
            status_code=400,
            detail="Необходимо указать причину удаления машины"
        )

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, brand, model, plate_number, is_deleted
                FROM vehicles
                WHERE id = %s
                """,
                (vehicle_id,)
            )
            vehicle = cursor.fetchone()

            if not vehicle:
                raise HTTPException(status_code=404, detail="Машина не найдена")

            if vehicle["is_deleted"]:
                raise HTTPException(status_code=400, detail="Машина уже удалена")

            cursor.execute(
                """
                SELECT COUNT(DISTINCT r.id) AS active_count
                FROM request_vehicles rv
                INNER JOIN requests r ON rv.request_id = r.id
                WHERE rv.vehicle_id = %s
                AND r.is_deleted = 0
                AND r.status IN ('NEW', 'IN_PROGRESS')
                """,
                (vehicle_id,)
            )
            result = cursor.fetchone()

            if result["active_count"] > 0:
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя удалить машину: по ней есть активные заявки"
                )

            cursor.execute(
                """
                UPDATE vehicles
                SET is_deleted = 1,
                    deleted_at = NOW(),
                    deleted_by = %s,
                    delete_reason_type = %s,
                    delete_reason = %s
                WHERE id = %s
                """,
                (
                    current_user["id"],
                    delete_reason_type,
                    delete_reason,
                    vehicle_id,
                )
            )

            connection.commit()

            return {
                "message": "Машина перемещена в корзину",
                "vehicle_id": vehicle_id,
                "delete_reason_type": delete_reason_type,
                "delete_reason": delete_reason,
            }

    except HTTPException:
        connection.rollback()
        raise
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@router.patch("/{vehicle_id}/restore")
def restore_vehicle(vehicle_id: int, current_user: dict = Depends(get_current_user)):
    """Восстановление машины из корзины."""
    if not can_restore_vehicle(current_user):
        raise HTTPException(status_code=403, detail="Недостаточно прав для восстановления машины")

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT 
                    v.id,
                    v.client_id,
                    v.vin,
                    v.is_deleted,
                    c.is_deleted AS client_is_deleted
                FROM vehicles v
                LEFT JOIN clients c ON v.client_id = c.id
                WHERE v.id = %s
                """,
                (vehicle_id,)
            )
            vehicle = cursor.fetchone()

            if not vehicle:
                raise HTTPException(status_code=404, detail="Машина не найдена")

            if not vehicle["is_deleted"]:
                raise HTTPException(status_code=400, detail="Машина не находится в корзине")

            if vehicle["client_id"] is None or vehicle["client_is_deleted"]:
                raise HTTPException(
                    status_code=400,
                    detail="Нельзя восстановить машину: клиент не найден или находится в корзине"
                )

            restore_vin = normalize_vin(vehicle.get("vin"))

            if restore_vin:
                cursor.execute(
                    """
                    SELECT
                        id,
                        client_id,
                        brand,
                        model,
                        plate_number
                    FROM vehicles
                    WHERE vin = %s
                    AND id != %s
                    AND is_deleted = 0
                    LIMIT 1
                    """,
                    (restore_vin, vehicle_id)
                )

                active_duplicate = cursor.fetchone()

                if active_duplicate:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Нельзя восстановить машину: VIN {restore_vin} уже используется "
                            f"активной машиной ID {active_duplicate['id']}"
                        )
                    )

            cursor.execute(
                """
                UPDATE vehicles
                SET vin = %s,
                    is_deleted = 0,
                    deleted_at = NULL,
                    deleted_by = NULL,
                    delete_reason_type = NULL,
                    delete_reason = NULL
                WHERE id = %s
                """,
                (restore_vin, vehicle_id)
            )

            connection.commit()

            return {
                "message": "Машина восстановлена",
                "vehicle_id": vehicle_id
            }

    except HTTPException:
        raise
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()