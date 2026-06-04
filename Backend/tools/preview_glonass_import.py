from __future__ import annotations

import re
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]

IMPORT_DIR = BASE_DIR / "import_data"
OUTPUT_DIR = BASE_DIR / "import_output"

CLIENTS_FILE = IMPORT_DIR / "Клиенты.xlsx"
USERS_FILE = IMPORT_DIR / "Пользователи.xlsx"
OBJECTS_FILE = IMPORT_DIR / "Список объектов.xlsx"

OUTPUT_FILE = OUTPUT_DIR / "glonass_import_preview.xlsx"

SOURCE_SYSTEM = "GLONASS_SOFT"

VIN_RE = re.compile(r"\b[A-HJ-NPR-Z0-9]{17}\b", re.IGNORECASE)


def normalize_text(value) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_key(value) -> str:
    return normalize_text(value).lower()


def normalize_phone(value) -> str:
    value = normalize_text(value)

    if not value:
        return ""

    return value


def read_sheet_as_dicts(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_text(h) for h in rows[0]]

    result = []

    for row in rows[1:]:
        item = {}

        for index, header in enumerate(headers):
            item[header] = row[index] if index < len(row) else None

        result.append(item)

    return result


def guess_client_type(client_name: str) -> str:
    name = normalize_text(client_name).upper()

    if name.startswith("ИП") or " ИП " in f" {name} ":
        return "IP"

    # Пока для импорта GlonassSoft почти всё считаем организациями.
    # Физлиц лучше потом отдельно проверить в preview.
    return "TOO"


def clean_email(value) -> str | None:
    email = normalize_text(value)

    if not email or "@" not in email:
        return None

    return email


def name_from_email(email: str | None) -> str | None:
    if not email or "@" not in email:
        return None

    login = email.split("@", 1)[0]
    login = re.sub(r"[._\-]+", " ", login)
    login = normalize_text(login)

    return login or None


def build_representative_name(user: dict | None, company_name: str) -> tuple[str, bool]:
    if user:
        first_name = normalize_text(user.get("Имя"))
        last_name = normalize_text(user.get("Фамилия"))

        fio = normalize_text(f"{last_name} {first_name}")

        if fio:
            return fio, False

        email = clean_email(user.get("Email")) or clean_email(user.get("Логин"))
        email_name = name_from_email(email)

        if email_name:
            return email_name, True

    return f"Представитель {company_name}", True


def build_placeholder_phone(index: int) -> str:
    # +7 + 10 цифр. Начинаем с +70000000001.
    return f"+7{index:010d}"


def extract_vin(custom_fields) -> str | None:
    text = normalize_text(custom_fields).upper()

    match = VIN_RE.search(text)

    if not match:
        return None

    return match.group(0).upper()


def clean_object_name(name: str) -> str:
    value = normalize_text(name)
    value = re.sub(r"\s*\(маяк\)\s*$", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s*\(.*?маяк.*?\)\s*$", "", value, flags=re.IGNORECASE)

    return normalize_text(value)


def extract_plate_from_object_name(object_name: str) -> str:
    cleaned = clean_object_name(object_name)

    # Часто имя выглядит как: BMW 535I 823BAY01
    # Берём последний токен, если он похож на госномер.
    parts = cleaned.split()

    if not parts:
        return "без ГРНЗ"

    candidate = parts[-1].upper()

    if re.fullmatch(r"[A-ZА-Я0-9]{5,10}", candidate):
        return candidate

    return cleaned


def parse_brand_model(object_name: str, model_object: str) -> tuple[str, str]:
    object_name = clean_object_name(object_name)
    model_object = normalize_text(model_object)

    bad_model_values = {
        "",
        "автомобиль",
        "на ходу. gps не работает",
        "auto",
    }

    if model_object.lower() not in bad_model_values:
        parts = model_object.split(maxsplit=1)

        if len(parts) == 1:
            return parts[0], ""

        return parts[0], parts[1]

    parts = object_name.split()

    if len(parts) >= 2:
        # Убираем последний токен, если он похож на госномер.
        maybe_plate = parts[-1].upper()

        if re.fullmatch(r"[A-ZА-Я0-9]{5,10}", maybe_plate):
            parts = parts[:-1]

    if len(parts) == 0:
        return "Не указано", "Не указано"

    if len(parts) == 1:
        return parts[0], "Не указано"

    return parts[0], " ".join(parts[1:])


def make_generated_vin(imei: str | None, fallback_index: int) -> str:
    imei = normalize_text(imei)

    if imei and re.fullmatch(r"\d{15}", imei):
        return f"GS{imei}"

    # 17 символов: GSOBJ + 12 цифр
    return f"GSOBJ{fallback_index:012d}"


def autofit_worksheet(ws):
    for column_cells in ws.columns:
        max_length = 0
        col_idx = column_cells[0].column

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        width = min(max(max_length + 2, 10), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    autofit_worksheet(ws)


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]):
    ws = wb.create_sheet(title)
    ws.append(headers)

    for row in rows:
        ws.append([row.get(header) for header in headers])

    style_sheet(ws)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    clients_raw = read_sheet_as_dicts(CLIENTS_FILE)
    users_raw = read_sheet_as_dicts(USERS_FILE)
    objects_raw = read_sheet_as_dicts(OBJECTS_FILE)

    warnings = []
    errors = []

    users_by_client = defaultdict(list)

    for user in users_raw:
        client_name = normalize_text(user.get("Клиент"))

        if client_name:
            users_by_client[normalize_key(client_name)].append(user)

    clients_by_name = {}

    for client in clients_raw:
        client_name = normalize_text(client.get("Имя"))

        if client_name:
            clients_by_name[normalize_key(client_name)] = client

    clients_normalized = []
    placeholder_phone_counter = 1

    for client in clients_raw:
        source_client_name = normalize_text(client.get("Имя"))
        source_parent_client_name = normalize_text(client.get("Клиент"))
        source_inn = normalize_text(client.get("ИНН"))

        if not source_client_name:
            errors.append({
                "type": "CLIENT_EMPTY_NAME",
                "message": "В Клиенты.xlsx найдена строка без Имя",
                "raw": str(client),
            })
            continue

        matched_users = users_by_client.get(normalize_key(source_client_name), [])
        selected_user = matched_users[0] if matched_users else None

        crm_email = None
        crm_phone = ""

        if selected_user:
            crm_email = clean_email(selected_user.get("Email")) or clean_email(selected_user.get("Логин"))
            crm_phone = normalize_phone(selected_user.get("Телефон"))

        phone_is_placeholder = False

        if not crm_phone:
            crm_phone = build_placeholder_phone(placeholder_phone_counter)
            placeholder_phone_counter += 1
            phone_is_placeholder = True

        crm_representative_name, name_is_placeholder = build_representative_name(
            selected_user,
            source_client_name,
        )

        clients_normalized.append({
            "source_system": SOURCE_SYSTEM,
            "source_client_name": source_client_name,
            "source_parent_client_name": source_parent_client_name,
            "source_inn": source_inn,
            "crm_type": guess_client_type(source_client_name),
            "crm_company_name": source_client_name,
            "crm_representative_name": crm_representative_name,
            "crm_phone": crm_phone,
            "crm_email": crm_email,
            "phone_is_placeholder": phone_is_placeholder,
            "name_is_placeholder": name_is_placeholder,
            "is_subclient": normalize_key(source_client_name) != normalize_key(source_parent_client_name),
            "matched_users_count": len(matched_users),
        })

    vehicles_map = {}
    devices_normalized = []
    generated_vin_counter = 1

    for obj in objects_raw:
        source_client_name = normalize_text(obj.get("Клиент"))
        object_name = normalize_text(obj.get("Имя объекта"))
        object_name_clean = clean_object_name(object_name)
        device_name = normalize_text(obj.get("Устройство"))
        imei = normalize_text(obj.get("IMEI"))
        custom_fields = normalize_text(obj.get("Произвольные поля"))
        model_object = normalize_text(obj.get("Модель объекта"))
        comment = normalize_text(obj.get("Комментарий"))

        if not source_client_name:
            errors.append({
                "type": "OBJECT_EMPTY_CLIENT",
                "message": "В Список объектов.xlsx найден объект без Клиент",
                "object_name": object_name,
                "imei": imei,
            })
            continue

        if normalize_key(source_client_name) not in clients_by_name:
            warnings.append({
                "type": "OBJECT_SKIPPED_CLIENT_NOT_FOUND",
                "message": "Объект пропущен: клиент объекта не найден в Клиенты.xlsx по колонке Имя",
                "source_client_name": source_client_name,
                "object_name": object_name,
                "imei": imei,
            })
            continue

        vin = extract_vin(custom_fields)
        vin_is_generated = False

        if not vin:
            vin = make_generated_vin(imei, generated_vin_counter)
            generated_vin_counter += 1
            vin_is_generated = True

        vehicle_key = (normalize_key(source_client_name), vin)

        brand, model = parse_brand_model(object_name_clean, model_object)
        plate_number = extract_plate_from_object_name(object_name_clean)

        if vehicle_key not in vehicles_map:
            vehicles_map[vehicle_key] = {
                "source_client_name": source_client_name,
                "object_name": object_name_clean,
                "plate_number": plate_number,
                "brand": brand,
                "model": model,
                "source_model_object": model_object,
                "generated_vin": vin,
                "vin_is_generated": vin_is_generated,
                "vehicle_type": "Легковая",
                "devices_count": 0,
                "imei_list": "",
            }

        vehicle = vehicles_map[vehicle_key]
        vehicle["devices_count"] += 1

        imei_values = [x for x in vehicle["imei_list"].split(", ") if x]

        if imei and imei not in imei_values:
            imei_values.append(imei)

        vehicle["imei_list"] = ", ".join(imei_values)

        devices_normalized.append({
            "source_client_name": source_client_name,
            "object_name": object_name,
            "object_name_clean": object_name_clean,
            "imei": imei,
            "device_name": device_name,
            "device_model": device_name,
            "linked_vin": vin,
            "vehicle_plate_number": plate_number,
            "status": "INSTALLED_PREVIEW_ONLY",
            "comment": comment,
        })

    vehicles_normalized = list(vehicles_map.values())

    # Проверка VIN, который встретился у разных клиентов.
    vin_to_clients = defaultdict(set)

    for vehicle in vehicles_normalized:
        vin_to_clients[vehicle["generated_vin"]].add(vehicle["source_client_name"])

    for vin, client_names in vin_to_clients.items():
        if len(client_names) > 1:
            warnings.append({
                "type": "VIN_USED_BY_MULTIPLE_CLIENTS",
                "message": "Один VIN встретился у нескольких клиентов",
                "vin": vin,
                "clients": ", ".join(sorted(client_names)),
            })

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_sheet(
        wb,
        "Clients_Normalized",
        [
            "source_system",
            "source_client_name",
            "source_parent_client_name",
            "source_inn",
            "crm_type",
            "crm_company_name",
            "crm_representative_name",
            "crm_phone",
            "crm_email",
            "phone_is_placeholder",
            "name_is_placeholder",
            "is_subclient",
            "matched_users_count",
        ],
        clients_normalized,
    )

    write_sheet(
        wb,
        "Vehicles_Normalized",
        [
            "source_client_name",
            "object_name",
            "plate_number",
            "brand",
            "model",
            "source_model_object",
            "generated_vin",
            "vin_is_generated",
            "vehicle_type",
            "devices_count",
            "imei_list",
        ],
        vehicles_normalized,
    )

    write_sheet(
        wb,
        "Devices_Normalized",
        [
            "source_client_name",
            "object_name",
            "object_name_clean",
            "imei",
            "device_name",
            "device_model",
            "linked_vin",
            "vehicle_plate_number",
            "status",
            "comment",
        ],
        devices_normalized,
    )

    write_sheet(
        wb,
        "Warnings",
        ["type", "message", "source_client_name", "object_name", "imei", "vin", "clients"],
        warnings,
    )

    write_sheet(
        wb,
        "Errors",
        ["type", "message", "raw", "object_name", "imei"],
        errors,
    )

    wb.save(OUTPUT_FILE)

    print("Preview created:")
    print(OUTPUT_FILE)
    print()
    print(f"Clients normalized: {len(clients_normalized)}")
    print(f"Vehicles normalized: {len(vehicles_normalized)}")
    print(f"Devices normalized: {len(devices_normalized)}")
    print(f"Warnings: {len(warnings)}")
    print(f"Errors: {len(errors)}")


if __name__ == "__main__":
    main()