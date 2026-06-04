from __future__ import annotations

import os
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
sys.path.append(str(BASE_DIR))

from app.database import get_connection

PREVIEW_FILE = BASE_DIR / "import_output" / "glonass_import_preview.xlsx"

SOURCE_SYSTEM = "GLONASS_SOFT"


def normalize_text(value) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_key(value) -> str:
    return normalize_text(value).lower()

def truncate_text(value, max_length: int) -> str:
    value = normalize_text(value)

    if len(value) <= max_length:
        return value

    return value[:max_length].rstrip()

def read_sheet(path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Лист {sheet_name} не найден в {path}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_text(h) for h in rows[0]]
    result = []

    for row in rows[1:]:
        item = {}

        for index, header in enumerate(headers):
            item[header] = row[index] if index < len(row) else None

        result.append(item)

    return result


def find_existing_client(cursor, source_client_name: str, company_name: str):
    """
    Приоритет:
    1. source_system + source_client_name
    2. company_name среди активных клиентов
    """
    cursor.execute(
        """
        SELECT id, company_name, name, phone, source_client_name
        FROM clients
        WHERE is_deleted = 0
          AND source_system = %s
          AND source_client_name = %s
        LIMIT 1
        """,
        (SOURCE_SYSTEM, source_client_name),
    )
    client = cursor.fetchone()

    if client:
        return client, "FOUND_BY_SOURCE"

    cursor.execute(
        """
        SELECT id, company_name, name, phone, source_client_name
        FROM clients
        WHERE is_deleted = 0
          AND LOWER(TRIM(company_name)) = LOWER(TRIM(%s))
        LIMIT 1
        """,
        (company_name,),
    )
    client = cursor.fetchone()

    if client:
        return client, "FOUND_BY_COMPANY_NAME"

    return None, None


def find_existing_vehicle(cursor, vin: str):
    cursor.execute(
        """
        SELECT id, client_id, brand, model, plate_number, vin
        FROM vehicles
        WHERE vin = %s
          AND is_deleted = 0
        LIMIT 1
        """,
        (vin,),
    )
    return cursor.fetchone()


def import_clients(cursor, clients_rows: list[dict], dry_run: bool):
    stats = {
        "created": 0,
        "existing_by_source": 0,
        "existing_by_company_name": 0,
        "errors": 0,
    }

    client_id_by_source_name = {}
    messages = []

    for index, row in enumerate(clients_rows, start=2):
        source_client_name = normalize_text(row.get("source_client_name"))
        source_parent_client_name = normalize_text(row.get("source_parent_client_name"))
        source_inn = normalize_text(row.get("source_inn"))

        crm_type = normalize_text(row.get("crm_type")) or "TOO"
        crm_company_name = normalize_text(row.get("crm_company_name"))
        crm_representative_name = normalize_text(row.get("crm_representative_name"))
        crm_phone = normalize_text(row.get("crm_phone"))
        crm_email = normalize_text(row.get("crm_email")) or None

        if not source_client_name or not crm_company_name:
            stats["errors"] += 1
            messages.append(
                f"[CLIENT row {index}] ERROR: пустой source_client_name/crm_company_name"
            )
            continue

        if not crm_representative_name:
            crm_representative_name = f"Представитель {crm_company_name}"

        if not crm_phone:
            stats["errors"] += 1
            messages.append(
                f"[CLIENT row {index}] ERROR: пустой crm_phone у {crm_company_name}"
            )
            continue

        existing_client, found_by = find_existing_client(
            cursor,
            source_client_name=source_client_name,
            company_name=crm_company_name,
        )

        if existing_client:
            client_id_by_source_name[normalize_key(source_client_name)] = existing_client["id"]

            if found_by == "FOUND_BY_SOURCE":
                stats["existing_by_source"] += 1
            else:
                stats["existing_by_company_name"] += 1

            messages.append(
                f"[CLIENT row {index}] SKIP: already exists id={existing_client['id']} "
                f"company={crm_company_name} reason={found_by}"
            )
            continue

        stats["created"] += 1

        if dry_run:
            fake_id = f"DRY_CLIENT_{stats['created']}"
            client_id_by_source_name[normalize_key(source_client_name)] = fake_id
            messages.append(
                f"[CLIENT row {index}] DRY CREATE: {crm_company_name}, phone={crm_phone}"
            )
            continue

        cursor.execute(
            """
            INSERT INTO clients (
                type,
                name,
                company_name,
                phone,
                email,
                source_system,
                source_client_name,
                source_parent_client_name,
                source_inn
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                crm_type,
                crm_representative_name,
                crm_company_name,
                crm_phone,
                crm_email,
                SOURCE_SYSTEM,
                source_client_name,
                source_parent_client_name or None,
                source_inn or None,
            ),
        )

        new_client_id = cursor.lastrowid
        client_id_by_source_name[normalize_key(source_client_name)] = new_client_id

        messages.append(
            f"[CLIENT row {index}] CREATE: id={new_client_id}, company={crm_company_name}"
        )

    return stats, client_id_by_source_name, messages


def import_vehicles(
    cursor,
    vehicles_rows: list[dict],
    client_id_by_source_name: dict,
    dry_run: bool,
):
    stats = {
        "created": 0,
        "existing_by_vin": 0,
        "client_not_found": 0,
        "errors": 0,
    }

    messages = []

    for index, row in enumerate(vehicles_rows, start=2):
        source_client_name = normalize_text(row.get("source_client_name"))
        client_key = normalize_key(source_client_name)

        client_id = client_id_by_source_name.get(client_key)

        if not client_id:
            stats["client_not_found"] += 1
            messages.append(
                f"[VEHICLE row {index}] ERROR: client not found for source_client_name={source_client_name}"
            )
            continue

        raw_plate_number = normalize_text(row.get("plate_number")) or "без ГРНЗ"

        brand = truncate_text(row.get("brand"), 100) or "Не указано"
        model = truncate_text(row.get("model"), 100) or "Не указано"
        plate_number = truncate_text(raw_plate_number, 50) or "без ГРНЗ"
        vin = truncate_text(row.get("generated_vin"), 100).upper()
        vehicle_type = truncate_text(row.get("vehicle_type"), 50) or "Легковая"

        if len(raw_plate_number) > 50:
            messages.append(
                f"[VEHICLE row {index}] WARNING: plate_number truncated "
                f"from '{raw_plate_number}' to '{plate_number}'"
            )

        if not vin:
            stats["errors"] += 1
            messages.append(
                f"[VEHICLE row {index}] ERROR: empty VIN, client={source_client_name}, object={row.get('object_name')}"
            )
            continue

        existing_vehicle = find_existing_vehicle(cursor, vin)

        if existing_vehicle:
            stats["existing_by_vin"] += 1
            messages.append(
                f"[VEHICLE row {index}] SKIP: vehicle already exists id={existing_vehicle['id']} vin={vin}"
            )
            continue

        stats["created"] += 1

        if dry_run:
            messages.append(
                f"[VEHICLE row {index}] DRY CREATE: client={source_client_name}, "
                f"{brand} {model}, plate={plate_number}, vin={vin}"
            )
            continue

        cursor.execute(
            """
            INSERT INTO vehicles (
                client_id,
                brand,
                model,
                plate_number,
                vin,
                year,
                type
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (
                client_id,
                brand,
                model,
                plate_number,
                vin,
                None,
                vehicle_type,
            ),
        )

        new_vehicle_id = cursor.lastrowid

        messages.append(
            f"[VEHICLE row {index}] CREATE: id={new_vehicle_id}, client_id={client_id}, vin={vin}"
        )

    return stats, messages


def print_stats(title: str, stats: dict):
    print()
    print(title)
    print("-" * len(title))

    for key, value in stats.items():
        print(f"{key}: {value}")


def main():
    if not PREVIEW_FILE.exists():
        raise RuntimeError(f"Preview file not found: {PREVIEW_FILE}")

    confirm = os.getenv("GLONASS_IMPORT_CONFIRM") == "YES"
    dry_run = not confirm

    print("GlonassSoft clients/vehicles import")
    print(f"Preview file: {PREVIEW_FILE}")
    print(f"Mode: {'REAL IMPORT' if confirm else 'DRY RUN'}")
    print()

    if dry_run:
        print("Чтобы выполнить реальный импорт, запусти:")
        print("set GLONASS_IMPORT_CONFIRM=YES")
        print("python tools/import_glonass_clients_vehicles.py")
        print()

    clients_rows = read_sheet(PREVIEW_FILE, "Clients_Normalized")
    vehicles_rows = read_sheet(PREVIEW_FILE, "Vehicles_Normalized")

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            client_stats, client_id_by_source_name, client_messages = import_clients(
                cursor=cursor,
                clients_rows=clients_rows,
                dry_run=dry_run,
            )

            vehicle_stats, vehicle_messages = import_vehicles(
                cursor=cursor,
                vehicles_rows=vehicles_rows,
                client_id_by_source_name=client_id_by_source_name,
                dry_run=dry_run,
            )

        if dry_run:
            connection.rollback()
            print("DRY RUN finished. No data was written.")
        else:
            connection.commit()
            print("REAL IMPORT finished. Data committed.")

        print_stats("Client stats", client_stats)
        print_stats("Vehicle stats", vehicle_stats)

        output_log = BASE_DIR / "import_output" / (
            "glonass_import_dry_run.log" if dry_run else "glonass_import_real.log"
        )

        with open(output_log, "w", encoding="utf-8") as f:
            f.write("CLIENTS\n")
            f.write("=" * 80 + "\n")
            f.write("\n".join(client_messages))
            f.write("\n\nVEHICLES\n")
            f.write("=" * 80 + "\n")
            f.write("\n".join(vehicle_messages))

        print()
        print(f"Log saved: {output_log}")

    except Exception as e:
        connection.rollback()
        print(f"IMPORT FAILED: {e}")
        raise

    finally:
        connection.close()


if __name__ == "__main__":
    main()